"""
raw_parser.py - Streaming parser for "Regular Demand Vs Collection" raw data.
Reads 268K+ rows via openpyxl read-only mode, aggregates by Branch > Officer,
and returns a JSON-serializable dict matching the frontend data contract.
"""

import io
from openpyxl import load_workbook


# Column name variations -> canonical key
COLUMN_MAP = {
    'branchname': 'branch_name', 'branch name': 'branch_name', 'branch': 'branch_name',
    'officer name': 'officer_name', 'officername': 'officer_name',
    'field officer': 'officer_name', 'fo name': 'officer_name',
    'officerid': 'officer_id', 'officer id': 'officer_id',
    'emp id': 'officer_id', 'empid': 'officer_id', 'employee id': 'officer_id',
    'emp code': 'officer_id',
    'account id': 'account_id', 'accountid': 'account_id',
    'account no': 'account_id', 'account number': 'account_id',
    'loan id': 'account_id', 'loanid': 'account_id',
    'client name': 'client_name', 'clientname': 'client_name',
    'customer name': 'client_name', 'borrower name': 'client_name',
    'product name': 'product', 'productname': 'product',
    'product': 'product', 'product type': 'product',
    'loan amount': 'loan_amount', 'loanamount': 'loan_amount',
    'loan amt': 'loan_amount', 'disbursed amount': 'loan_amount',
    'sanctioned amount': 'loan_amount',
    'dpd days': 'dpd_days', 'dpd': 'dpd_days',
    'days past due': 'dpd_days', 'dpddays': 'dpd_days',
    'dpd group': 'dpd_group', 'dpd bucket': 'dpd_group',
    'dpd range': 'dpd_group', 'dpdgroup': 'dpd_group', 'dpd band': 'dpd_group',
    'current loan status': 'status', 'loan status': 'status',
    'status': 'status', 'account status': 'status',
    'regular demand': 'regular_demand', 'regulardemand': 'regular_demand',
    'reg demand': 'regular_demand',
    'cumulative demand': 'cumulative_demand', 'cumulativedemand': 'cumulative_demand',
    'cum demand': 'cumulative_demand', 'total demand': 'cumulative_demand',
    'collection': 'collection', 'total collection': 'collection',
    'collection amount': 'collection',
    'partial amount': 'partial_amount', 'partialamount': 'partial_amount',
    'partial': 'partial_amount', 'partial amt': 'partial_amount',
}


def _safe_num(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val):
    if val is None:
        return ''
    return str(val).strip()


def _dpd_bucket(dpd_group_str, dpd_days):
    g = str(dpd_group_str).lower().strip() if dpd_group_str else ''
    if '0' in g and ('day' in g or 'dpd' in g or g == '0'):
        return '0_days'
    if '1' in g and '30' in g:
        return '1_30'
    if '31' in g and '60' in g:
        return '31_60'
    if '61' in g and '90' in g:
        return '61_90'
    if '90' in g and ('+' in g or 'above' in g or 'plus' in g):
        return '90_plus'
    if '>90' in g or '91' in g:
        return '90_plus'
    d = _safe_num(dpd_days)
    if d <= 0:
        return '0_days'
    if d <= 30:
        return '1_30'
    if d <= 60:
        return '31_60'
    if d <= 90:
        return '61_90'
    return '90_plus'


def _make_totals():
    return {
        'accounts': 0,
        'regular_demand': 0.0,
        'cumulative_demand': 0.0,
        'collection': 0.0,
        'collection_pct': 0.0,
        'dpd_breakdown': {
            '0_days': 0, '1_30': 0, '31_60': 0, '61_90': 0, '90_plus': 0
        }
    }


def _finalize_totals(totals):
    if totals['regular_demand'] > 0:
        totals['collection_pct'] = round(
            (totals['collection'] / totals['regular_demand']) * 100, 1
        )
    else:
        totals['collection_pct'] = 0.0


def parse_raw_file(file_bytes, include_accounts=True):
    """
    Parse the raw Excel file. Returns dict matching the JSON data contract.

    Args:
        file_bytes: bytes of the .xlsx file
        include_accounts: if False, skip account-level detail (saves memory)
    """
    wb = load_workbook(
        filename=io.BytesIO(file_bytes),
        read_only=True,
        data_only=True
    )

    ws = wb.active

    # --- Find header row and map columns ---
    col_indices = {}
    header_row = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row is None:
            continue
        for col_idx, cell_val in enumerate(row):
            if cell_val is None:
                continue
            normalized = str(cell_val).strip().lower()
            if normalized in COLUMN_MAP:
                canonical = COLUMN_MAP[normalized]
                if canonical not in col_indices:
                    col_indices[canonical] = col_idx
                    header_row = row_idx
        if len(col_indices) >= 5:
            break

    if not header_row or 'branch_name' not in col_indices:
        wb.close()
        raise ValueError(
            'Could not find expected columns in the raw data file. '
            'Expected at least: BranchName, Officer Name. '
            'Found: ' + ', '.join(col_indices.keys())
        )

    # --- Stream all data rows ---
    branches = {}
    total_rows = 0

    def get(row, key):
        idx = col_indices.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None:
            continue

        branch = _safe_str(get(row, 'branch_name'))
        if not branch:
            continue

        total_rows += 1
        officer_name = _safe_str(get(row, 'officer_name')) or 'Unknown'
        officer_id = _safe_str(get(row, 'officer_id')) or ''

        regular_demand = _safe_num(get(row, 'regular_demand'))
        cumulative_demand = _safe_num(get(row, 'cumulative_demand'))
        collection = _safe_num(get(row, 'collection'))
        dpd_group = get(row, 'dpd_group')
        dpd_days = get(row, 'dpd_days')
        bucket = _dpd_bucket(dpd_group, dpd_days)

        # Ensure branch exists
        if branch not in branches:
            branches[branch] = {'totals': _make_totals(), 'officers': {}}
        br = branches[branch]

        # Ensure officer exists
        if officer_name not in br['officers']:
            br['officers'][officer_name] = {
                'officer_id': officer_id,
                'totals': _make_totals(),
                'accounts': [] if include_accounts else None
            }
        off = br['officers'][officer_name]

        if officer_id and not off['officer_id']:
            off['officer_id'] = officer_id

        # Accumulate officer totals
        off['totals']['accounts'] += 1
        off['totals']['regular_demand'] += regular_demand
        off['totals']['cumulative_demand'] += cumulative_demand
        off['totals']['collection'] += collection
        off['totals']['dpd_breakdown'][bucket] += 1

        # Accumulate branch totals
        br['totals']['accounts'] += 1
        br['totals']['regular_demand'] += regular_demand
        br['totals']['cumulative_demand'] += cumulative_demand
        br['totals']['collection'] += collection
        br['totals']['dpd_breakdown'][bucket] += 1

        # Account detail
        if include_accounts and off['accounts'] is not None:
            off['accounts'].append({
                'account_id': _safe_str(get(row, 'account_id')),
                'client_name': _safe_str(get(row, 'client_name')),
                'product': _safe_str(get(row, 'product')),
                'loan_amount': _safe_num(get(row, 'loan_amount')),
                'dpd_days': _safe_num(dpd_days) if dpd_days is not None else 0,
                'dpd_group': _safe_str(dpd_group),
                'status': _safe_str(get(row, 'status')),
                'regular_demand': regular_demand,
                'collection': collection,
                'partial_amount': _safe_str(get(row, 'partial_amount'))
            })

    wb.close()

    # Finalize collection percentages
    total_officers = 0
    for br in branches.values():
        _finalize_totals(br['totals'])
        for off in br['officers'].values():
            _finalize_totals(off['totals'])
            total_officers += 1

    return {
        'status': 'ok',
        'meta': {
            'total_rows': total_rows,
            'total_branches': len(branches),
            'total_officers': total_officers
        },
        'branches': branches
    }
